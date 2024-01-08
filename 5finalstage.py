from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

product_codes = ["605.260.85", "605.389.55"]
options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])

driver = webdriver.Chrome("C:/Users/Dell User/Desktop/chromedriver.exe", options=options)

desktop_path = "C:/Users/Dell User/Desktop/"
excel_file_path = desktop_path + "urun_bilgileri.xlsx"

try:
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ürün Bilgileri"
    sheet.append(["Ürün Kodu", "Fiyat", "İnternet Stok Durumu", "Kartal Stok Durumu", "Ümraniye Stok Durumu", "Resim URL"])
    bold_font = Font(bold=True)  # bu kısmı columnları kalın yapmak için ekledim
    for cell in sheet[1]:
        cell.font = bold_font

green_fill = PatternFill(start_color="CDDC39", end_color="CDDC39", fill_type="solid")
red_fill = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
yellow_fill = PatternFill(start_color="FFD600", end_color="FFD600", fill_type="solid")

for code in product_codes:
    url = "https://www.ikea.com.tr/"
    driver.get(url)
    time.sleep(1)
    driver.delete_all_cookies()
    time.sleep(1)
    search_box = driver.find_element(By.CSS_SELECTOR,
                                     '#ctl00_ctrlHeader_divSearchBox > div.search-wrapper > div > input')
    search_box.click()
    time.sleep(1)
    search_box.send_keys(code)
    search_box.send_keys(Keys.ENTER)
    time.sleep(1)

    try:
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_divPrice"]/span/span'))
        )
        price = element.text
    except:
        price = "Fiyat bulunamadı"

    try:
        mgzstok = driver.find_element(By.CSS_SELECTOR,
                                       '#aspnetForm > div.subpage-content > div > div.product-detail-wrapper > div:nth-child(1) > div.col-lg-4.col-md-5.col-xs-12 > div > a > span:nth-child(2)')
        driver.execute_script("arguments[0].scrollIntoView();", mgzstok)

        time.sleep(1)

        driver.execute_script("arguments[0].click();", mgzstok)
        time.sleep(1)
        stock_status = driver.find_element(By.XPATH,
                                           '//*[@id="check-stock-modal"]/div[2]/div[8]/div[2]/span').text
    except:
        stock_status = "İnternet stok bulunamadı"

    try:
        kartal_stok = driver.find_element(By.XPATH, '//*[@id="check-stock-modal"]/div[2]/div[6]/div[2]').text
    except:
        kartal_stok = "Kartal stok bulunamadı"

    try:
        umraniye_stok = driver.find_element(By.XPATH, '//*[@id="check-stock-modal"]/div[2]/div[7]/div[2]').text
    except:
        umraniye_stok = "Ümraniye stok bulunamadı"

    try:
        image_url = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_rptImages_ctl00_image"]').get_attribute('src')
    except:
        image_url = "Resim bulunamadı"

    sheet.append([code, price, stock_status, kartal_stok, umraniye_stok, image_url])

    if "Stokta Var" in stock_status:
        sheet.cell(row=sheet.max_row, column=3).fill = green_fill
    elif "Stokta Yok" in stock_status:
        sheet.cell(row=sheet.max_row, column=3).fill = red_fill
    elif "Kritik Stok" in stock_status:
        sheet.cell(row=sheet.max_row, column=3).fill = yellow_fill

    if "Stokta Var" in kartal_stok:
        sheet.cell(row=sheet.max_row, column=4).fill = green_fill
    elif "Stokta Yok" in kartal_stok:
        sheet.cell(row=sheet.max_row, column=4).fill = red_fill
    elif "Kritik Stok" in kartal_stok:
        sheet.cell(row=sheet.max_row, column=4).fill = yellow_fill

    if "Stokta Var" in umraniye_stok:
        sheet.cell(row=sheet.max_row, column=5).fill = green_fill
    elif "Stokta Yok" in umraniye_stok:
        sheet.cell(row=sheet.max_row, column=5).fill = red_fill
    elif "Kritik Stok" in umraniye_stok:
        sheet.cell(row=sheet.max_row, column=5).fill = yellow_fill

    for column in sheet.columns:
        max_uzunluk = 0
        column = [cell for cell in column]
        try:
            max_uzunluk = max(len(str(cell.value)) for cell in column)
            yeni_uzunluk = (max_uzunluk)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = yeni_uzunluk
        except:
            pass

workbook.save(excel_file_path)
driver.quit()
